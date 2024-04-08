import argparse
import shutil
import typing
import uuid
import openpyxl
import openpyxl.worksheet.worksheet
from cramerapis import CramerAPIs
from factoryproduct import FactoryProductConfiguration
from debug import DBG, set_debug_level


class TAFEntry:
    ''' A single row in the Excel Sheet for TAF '''
    product: str
    api: str
    send_verify: str
    request_response: str
    paramtype: str
    paramname: str
    paramvalue: str
    additional_attribute: str

    def __init__(self, product, api, send_verify, request_response, paramtype, paramname, paramvalue, add_attribute = None):
        self.product = product
        self.api = api
        self.send_verify = send_verify
        self.request_response = request_response
        self.paramtype = paramtype
        self.paramname = paramname
        self.paramvalue = paramvalue
        self.additional_attribute = add_attribute

    def __str__(self):
        return "{:<2s}  {:<6s}  {:<6s}  {:<16s}  {:<19s}  {:<50s}  {:<40s}  {}".format(
            self.product, self.api, self.send_verify, self.request_response, self.paramtype,
            self.paramname, self.paramvalue, self.additional_attribute)
        

class TAFTestCase:
    ''' Definition of a TAF Testcase '''
    caseid: str
    name: str
    request_entries: typing.List[TAFEntry]
    response_entries: typing.List[TAFEntry]
    task_entries: typing.List[TAFEntry]
    oh_entries: typing.List[TAFEntry]
    priority: str
    precondition: str
    objective: str
    expected_result: str
    comment: str
    ref_pre: str

    def __init__(self, caseid, name, *, priority=None, precondition=None, objective=None, expected_result=None, comment=None, ref_pre=None):
        self.caseid = caseid
        self.name = name
        self.request_entries = [TAFEntry("IL", "SOA WS", "Setup", "AsyncReceiver", "", "TimeOut", r"${TIMEOUT}")]
        self.response_entries = []
        self.task_entries = []
        self.oh_entries = []
        self.priority = priority
        self.precondition = precondition
        self.objective = objective
        self.expected_result = expected_result
        self.comment = comment
        self.ref_pre = ref_pre

#    def append(self, entry: TAFEntry):
#        self.entries.append(entry)

    def print(self):
        for e in self.request_entries + self.oh_entries + self.response_entries + self.task_entries:
            print(e)

class TAFExcel:
    ''' Class for the Excel Sheet holding the TAF Test Cases. '''
    filename: str
    wb: openpyxl.Workbook
    sheet: openpyxl.worksheet.worksheet.Worksheet
    nextrow: int


    def __init__(self, filename, template="C:/Users/reuss/Documents/Sunrise/OSS Modernization/TND/TAF/TAF_Template.xlsx"):
        self.filename = filename
        shutil.copy(template, filename)
        self.wb = openpyxl.load_workbook(filename)
        self.sheet = self.wb["Test cases"]
        self.nextrow = 3

    def add_test_case(self, tc : TAFTestCase):
        startrow = self.nextrow
        sheet = self.sheet
        sheet.cell(row=self.nextrow, column=1).value = tc.caseid
        sheet.cell(row=self.nextrow, column=2).value = tc.name
        sheet.cell(row=self.nextrow, column=3).value = tc.caseid + r"-${RUN_ID}"
        sheet.cell(row=self.nextrow, column=12).value = tc.priority if tc.priority else ""
        sheet.cell(row=self.nextrow, column=13).value = tc.precondition if tc.precondition else ""
        sheet.cell(row=self.nextrow, column=14).value = tc.objective if tc.objective else ""
        sheet.cell(row=self.nextrow, column=15).value = tc.expected_result if tc.expected_result else ""
        sheet.cell(row=self.nextrow, column=16).value = tc.comment if tc.comment else ""
        sheet.cell(row=self.nextrow, column=17).value = tc.ref_pre if tc.ref_pre else ""
        for entry in tc.request_entries + tc.oh_entries + tc.response_entries + tc.task_entries:
            sheet.cell(row=self.nextrow, column=4).value = entry.product
            sheet.cell(row=self.nextrow, column=5).value = entry.api
            sheet.cell(row=self.nextrow, column=6).value = entry.send_verify
            sheet.cell(row=self.nextrow, column=7).value = entry.request_response
            sheet.cell(row=self.nextrow, column=8).value = entry.paramtype
            sheet.cell(row=self.nextrow, column=9).value = entry.paramname
            sheet.cell(row=self.nextrow, column=10).value = entry.paramvalue
            if entry.additional_attribute is not None:
                sheet.cell(row=self.nextrow, column=11).value = entry.additional_attribute
            self.nextrow += 1
        for col in [1,2,3,12,13,14,15,16,17]:
            sheet.merge_cells(start_row=startrow, start_column=col, end_row=self.nextrow-1, end_column=col)
            sheet.cell(row=startrow, column=col).alignment = openpyxl.styles.alignment.Alignment(vertical='top', wrapText=True)

    def save(self):
        self.wb.save(self.filename)

    def close(self):
        self.save()
        self.wb.close()


class RequestHeaderEntry(TAFEntry):
    def __init__(self, name, value):
        super().__init__("IL", "SOA WS", "Send", "CreateRequest", "RequestHeader", name, value)

class RequestParameterEntry(TAFEntry):
    def __init__(self, name, value):
        super().__init__("IL", "SOA WS", "Send", "CreateRequest", "RequestParameter", name, value)

class ResponseHeaderEntry(TAFEntry):
    def __init__(self, name, value):
        super().__init__("IL", "SOA WS", "Verify", "Response", "ResponseHeader", name, value, "RESP")

class ResponseParameterEntry(TAFEntry):
    def __init__(self, name, value):
        super().__init__("IL", "SOA WS", "Verify", "Response", "ResponseParameter", name, value)

class TaskRequestEntry(TAFEntry):
    def __init__(self, name, value, taskid):
        super().__init__("IL", "NEI", "Verify", "create", "TaskRequest", name, value, "TASK={}".format(taskid))

class TaskResponseEntry(TAFEntry):
    def __init__(self, name, value, taskid):
        super().__init__("IL", "NEI", "Verify", "create", "TaskResponse", name, value, "TASK={}".format(taskid))

class GlobalVariableEntry(TAFEntry):
    def __init__(self, varname, taskid=None):
        super().__init__("IL", "NEI" if taskid else "SOA WS", "Robot", "Keyword", "Set Global Variable", r'${'+varname+r'}', "", "Task={}".format(taskid) if taskid else "")

class OrderHubSendTask(TAFEntry):
    def __init__(self, paramname, paramvalue, taskid, *, paramtype=""):
        super().__init__("OH", "REST", "Send", "OHTicket", paramtype, paramname, paramvalue, "TASK={}".format(taskid))

class OrderHubVerifyTask(TAFEntry):
    def __init__(self, paramname, paramvalue, taskid):
        super().__init__("OH", "REST", "Verify", "OHTicket", "", paramname, paramvalue, "TASK={}".format(taskid))

def createRequestEntries(tc : TAFTestCase, config : FactoryProductConfiguration, use_example_values=False, pause_after_prepare=True):
    tc.request_entries.append(RequestHeaderEntry("NeType", "CDFF"))
    tc.request_entries.append(RequestHeaderEntry("MaxReqTime", "86400"))
    for name, value in [ ("IL_REQ_GROUP", "TND"),
                         ("ORDER_TYPE", "Connect"),
                         ("ORDER_EXTERNAL_ORDER_ID", r"${ORDER_EXTERNAL_ORDER_ID}"),
                         ("LINE_1_ORDER_LINE_ID", "OL_1"),
                         ("LINE_1_NAME", "FACTORY_PRODUCT_{}".format(config.factoryProductName)),
                         ("LINE_1_ACTION", "Create")  ]:
        tc.request_entries.append(RequestParameterEntry(name, value))
    for name, value in [ ("PAUSE_AFTER_PREPARE", "true" if pause_after_prepare else "false")]:
        tc.request_entries.append(RequestParameterEntry("LINE_1_PS_PARAM_{}".format(name), value))
    for param in config.input_params:
        if use_example_values:
            tc.request_entries.append(RequestParameterEntry("LINE_1_PS_PARAM_{}".format(param.name), param.examplevalue))
        else:
            tc.request_entries.append(RequestParameterEntry("LINE_1_PS_PARAM_{}".format(param.name), r"${"+param.name+r"}"))
    tc.response_entries.append(ResponseHeaderEntry("Status", "9"))

def createSuborderTaskEntries(tc : TAFTestCase, config : FactoryProductConfiguration, taskid : int, *, cfsname=None, component=None):
    for name, value in [ ("NE_TYPE", "SOAP"),
                         ("ACTION", "create"),
                         ("PARAM_NE_TYPE", "CDFF"),
                         ("SEND_IL_REQ_GROUP", "TND"),
                         ("SEND_LINE_1_NAME", "TECHPROD_CRAMER_FACT_PROD_{}".format(config.factoryProductName)),
                         ("SEND_LINE_1_ACTION", "Create")  ]:
        tc.task_entries.append(TaskRequestEntry(name, value, taskid))
    if cfsname is not None:
        tc.task_entries.append(TaskRequestEntry("SEND_LINE_1_CFS_NAME", cfsname, taskid))
        if component is not None:
            tc.task_entries.append(TaskRequestEntry("SEND_LINE_1_COMPONENT_NAME", component, taskid))

    for param in config.input_params:
        tc.task_entries.append(TaskRequestEntry("SEND_LINE_1_PS_PARAM_{}".format(param.name), r"${"+param.name+r"}", taskid))
    tc.task_entries.append(TaskResponseEntry("SMESSAGE_ID", "7", taskid))
    for param in config.cramer_output_params:
        tc.task_entries.append(TaskResponseEntry("LINE_1_{}".format(param.name), "RF_VAR={}".format(param.name), taskid))
        tc.task_entries.append(GlobalVariableEntry(param.name, taskid))

def createStablenetTaskEntries(tc: TAFTestCase, config: FactoryProductConfiguration, taskid: int, *, cfsname=None, component=None):
    for name, value in [ ("NE_TYPE", "STABLENETTND"),
                         ("ACTION", "process"),
                         ("CFS_NAME", cfsname if cfsname is not None else "N/A"),
                         ("COMPONENT_NAME", component if component is not None else "N/A"),
                         ("FACTORY_PRODUCT_NAME", config.factoryProductName),
                         ("TARGET_DEVICE", r"${NETWORK_ELEMENT_NAME}"),
                         ("RFS_NAME", r"${"+config.factoryProductName+r"_RFS_NAME}")]:
        tc.task_entries.append(TaskRequestEntry(name, value, taskid))
    for i, param in enumerate(config.stablenet_params[0], 1):
        tc.task_entries.append(TaskRequestEntry("PARAMETER_NAME_{}".format(i), param, taskid))
        tc.task_entries.append(TaskRequestEntry("PARAMETER_VALUE_{}".format(i), r"${"+param+r"}", taskid))
    tc.task_entries.append(TaskResponseEntry("CODE", "0", taskid))
    tc.task_entries.append(TaskResponseEntry("SMESSAGE_ID", "0", taskid))

def createBasicCramerAPITaskEntries(sheet: TAFTestCase, apiname: str, taskid: int, *, cfsname=None, component=None):
    sheet.task_entries.append(TaskRequestEntry("NE_TYPE", "CRAMERREST", taskid))
    sheet.task_entries.append(TaskRequestEntry("ACTION", "executeQuery", taskid))
    sheet.task_entries.append(TaskRequestEntry("QUERY_ID", apiname, taskid))
    sheet.task_entries.append(TaskResponseEntry("SMESSAGE_ID", "0", taskid))

def createGenericTaskEntries(tc: TAFTestCase, netype: str, taskid: int, *, cfsname=None, component=None, action=None, smessageid=0):
    tc.task_entries.append(TaskRequestEntry("NE_TYPE", netype, taskid))
    if action:
        tc.task_entries.append(TaskRequestEntry("ACTION", action, taskid))
    if smessageid is not None:
        tc.task_entries.append(TaskResponseEntry("SMESSAGE_ID", str(smessageid), taskid))

def createOHTaskEntries(tc: TAFTestCase, config: FactoryProductConfiguration, orderhub_queue: str, taskid: int):
    tc.oh_entries.append(OrderHubSendTask("workqueue", orderhub_queue, taskid))
    tc.oh_entries.append(OrderHubSendTask("RESOLUTION", "ADD", taskid, paramtype="TaskParameter"))
    tc.oh_entries.append(OrderHubSendTask("status", "COMPLETE", taskid))
    tc.oh_entries.append(OrderHubVerifyTask("workqueue", orderhub_queue, taskid))
    tc.oh_entries.append(OrderHubVerifyTask("status", "closed", taskid))


def create_myvariables(filename: str, config: FactoryProductConfiguration):
    with open(filename, "w") as fp:
        fp.write('TIMEOUT = "900"\n\n')
        fp.write('ORDER_EXTERNAL_ORDER_ID = "{}"\n\n'.format(uuid.uuid4()))
        for param in config.input_params:
            fp.write('{} = "{}"\n'.format(param.name, param.examplevalue))


def addFactoryProductValidatePhase(tc: TAFTestCase, config: FactoryProductConfiguration, taskid: int, *, cfsname=None, component=None):
    # Cramer Validations
    for valtask in config.cramer_validations:
        createBasicCramerAPITaskEntries(tc, valtask.api, taskid)
        taskid += 1
    return taskid

def addFactoryProductPreparePhase(tc: TAFTestCase, config: FactoryProductConfiguration, taskid: int, *,
                                  cfsname=None, component=None, with_confirmation=True, with_reuse=False):
    # Acquire Lock
    createGenericTaskEntries(tc, "DB", taskid, action="ACQUIRE_LOCK", smessageid=None)
    taskid += 1
    # Identify
    identifyApi = CramerAPIs[config.factoryProductName][1]
    if identifyApi is not None:
        createBasicCramerAPITaskEntries(tc, identifyApi, taskid)
        tc.task_entries.append(TaskResponseEntry("SERVICE_FOUND", "true" if with_reuse else "false", taskid)) # No reuse
        taskid += 1
    if not with_reuse:
        # Suborder
        createSuborderTaskEntries(tc, config, taskid, cfsname=cfsname, component=component)
        taskid += 1
    elif cfsname:
        # Relate Services
        createBasicCramerAPITaskEntries(tc, "relateToService", taskid)
        taskid += 1
    if with_confirmation:
        # Release Lock
        createGenericTaskEntries(tc, "DB", taskid, action="RELEASE_LOCK", smessageid=None)
        taskid += 1
    return taskid

def addFactoryProductProvisionPhase(tc: TAFTestCase, config: FactoryProductConfiguration, taskid: int, *,
                                    cfsname=None, component=None, with_confirmation=True, with_reuse=False):
    if with_confirmation:
        # Order Hub
        createGenericTaskEntries(tc, "ORDERHUB", taskid)
        createOHTaskEntries(tc, config, "TND Confirmations", taskid)
        taskid += 1
        # Acquire Lock
        createGenericTaskEntries(tc, "DB", taskid, action="ACQUIRE_LOCK", smessageid=None)
        taskid += 1
    # Get PS from Cramer
    createBasicCramerAPITaskEntries(tc, "getServiceDetails", taskid)
    tc.task_entries.append(TaskResponseEntry("PROVISION_STATUS", "SERVICE - READY FOR ACTIVATION", taskid)) # No reuse
    taskid += 1
    # Stablenet
    createStablenetTaskEntries(tc, config, taskid)
    taskid += 1
    # Update PS
    createBasicCramerAPITaskEntries(tc, "updateServiceProvisionStatus", taskid)
    taskid += 1
    return taskid
    
def addFactoryProductFinalizePhase(tc: TAFTestCase, config: FactoryProductConfiguration, taskid: int, *, cfsname=None, component=None):
    # Release Lock
    createGenericTaskEntries(tc, "DB", taskid, action="RELEASE_LOCK", smessageid=None)
    taskid += 1
    return taskid
    

def createFactoryProductCreateTestCase(config: FactoryProductConfiguration, *, with_confirmation=True, with_reuse=False):
    factoryProduct = config.factoryProductName
    tc = TAFTestCase("Create_{}".format(factoryProduct),  "Create {} (No Reuse)".format(factoryProduct),
                     precondition="None", expected_result="Successful execution",
                     comment="No reuse of service, {} confirmation".format("with" if with_confirmation else "without"))
    createRequestEntries(tc, config, pause_after_prepare=with_confirmation)
    taskid = 2
    taskid = addFactoryProductValidatePhase(tc, config, taskid)
    # Create TND Context
    createBasicCramerAPITaskEntries(tc, "createTNDContext", taskid)
    taskid += 1
    taskid = addFactoryProductPreparePhase(tc, config, taskid, with_confirmation=with_confirmation, with_reuse=with_reuse)
    taskid = addFactoryProductProvisionPhase(tc, config, taskid, with_confirmation=with_confirmation, with_reuse=with_reuse)
    taskid = addFactoryProductFinalizePhase(tc, config, taskid)
    # Apply Context
    createBasicCramerAPITaskEntries(tc, "applyTNDContext", taskid)
    taskid += 1
    return tc


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('-D', dest='dbglevel', action='store', default=0,  help='Print Debug information')
    parser.add_argument('-o', dest='outfilename', action='store',          help='Output filename')
    parser.add_argument('filename', metavar='<FILENAME>',                  help='JSON input files')

    args=parser.parse_args()
    set_debug_level(args.dbglevel)


    DBG(10, "Reading json file '{}'".format(args.filename))
    with open(args.filename, "r") as fp:
        config = FactoryProductConfiguration.from_file(fp)
    tc1 = createFactoryProductCreateTestCase(config)
    #tc1.print()
    tc2 = createFactoryProductCreateTestCase(config, with_confirmation=False)

    outfilename = args.outfilename if args.outfilename is not None else "TAF_{}.xlsx".format(config.factoryProductName)
    DBG(10, "Creating Excel File {}".format(outfilename))
    xls = TAFExcel(outfilename)
    xls.add_test_case(tc1)
    xls.add_test_case(tc2)
    xls.close()
    create_myvariables("my_variables.py", config)



