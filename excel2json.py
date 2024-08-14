#!/usr/bin/python3
import argparse
import re
from zipfile import ZipFile
from collections import namedtuple
from openpyxl import load_workbook

from factoryproduct import Param, FactoryProductConfiguration
from debug import DBG, set_debug_level

FP_DEPENDENCIES = {"ELAN_SAP": ("ELAN_CORE", "ELAN_CORE_RFS_NAME"),
                   "ELINE_SAP": ("ELINE_CORE", "ELINE_CORE_RFS_NAME"),
                   "IPVPN_SAP": ("IPVPN_CORE", "IPVPN_CORE_RFS_NAME")}

def read_data_from_excel(xlfile, tab):
    ''' Read and return the data from the correct sheet in the excel file.'''
    PRODNAME_CELL = 'B1'
    VERSION_CELL = 'B4'

    DBG(10, "Reading data from excel file '{}'".format(xlfile))
    wb = load_workbook(xlfile, data_only=True)
    if tab is None:
        sheet = wb.active
        tab = sheet.title
        print("WARNING: No parameter given for <TABNAME>. Reading active tab '{}'".format(tab))
    else:
        DBG(10, "Reading tab '{}'".format(tab))
        try:
            sheet = wb[tab]
        except KeyError:
            tabs = ", ".join("'"+t+"'" for t in wb.sheetnames)
            raise KeyError("Invalid tab name '{}'. Available tabs are: {}".format(tab, tabs))

    prodname = sheet[PRODNAME_CELL].value
    if not prodname:
        raise ValueError("Unable to read product name from excel {}, tab {}, cell {}".format(xlfile, tab, PRODNAME_CELL))
    version = sheet[VERSION_CELL].value

    inparams, crameroutparams, stablenetparams, keyparams = read_params_from_sheet(sheet, prodname)
    # Add prerequisite parameters to Stablenet
    if prodname in FP_DEPENDENCIES:
        depprod = FP_DEPENDENCIES[prodname][0]
        try:
            sheet = wb[depprod]
        except KeyError:
            raise KeyError("Unable to open tab '{}' for prerequisite product.".format(depprod))
        _, _, additional_stablenetparams, _ = read_params_from_sheet(sheet, depprod)
        stablenetparams = tuple(stablenetparams[i] + ["{}_{}".format(depprod, pname) for pname in additional_stablenetparams[i]] for i in range(2))

    return prodname, version, inparams, crameroutparams, stablenetparams, keyparams

def read_params_from_sheet(sheet, prodname):
    ''' Read data from an excel sheet '''
    STARTROW = 7
    COLUMNS = {'TECHNAME': 1, 'DESC': 2, 'VALUETYPE': 3, 'TYPEDETAILS': 4, 'OM': 5,
               'EXAMPLE_VALUE': 6, 'CRAMERSTORAGE': 7, 'JSONNAME': 8, 'STABLENET': 9,
               'MODIFY': 10, 'PARAMTYPE': 11, 'ACADEFAULT': 13 }
    COLS_TO_CHECK_FOR_MAPPING = [14 ,16, 18, 20]
    # ACTION_CELL = 'B3'

    inparams = []
    crameroutparams = []
    stablenetparams = ([],[])
    stablenetblacklist = ["NETWORK_ELEMENT_NAME", '{}_RFS_NAME'.format(prodname), "EVPN_EVI_RANGE"]

    description_overrides = {"CUST_SNIPPET_NAMES": "Custom configuration snippet(s) for configuring the {} service".format(prodname)}

    keyparams = []
    for row in range(STARTROW, 300):
        cellvalues = [sheet.cell(row=row, column=col).value for col in [COLUMNS[s]
            for s in COLUMNS.keys()]]
        techname, desc, valuetype, typedetails, mo, example, cramerstorage, jsonname, stablenet, modify, paramtype, acadefault = \
            (x.strip() if isinstance(x, str) else x for x in cellvalues)
        if techname == "Version": # avoid reading version history
            break
        if techname is not None and techname.startswith("#"):
            continue
        if sheet.cell(row=row, column=1).font.strike:
            DBG(20, "Ignoring parameter {} because of strikethough formatting".format(techname))
            continue
        if techname and valuetype and mo and paramtype:
            # Check for arrays
            m = re.match(r'(.*)\[\d+\.\.(\d+)\]', valuetype)
            if m is not None:
                valuetype = m.group(1)
                maxoccurs = int(m.group(2))
            else:
                maxoccurs = None
            mapped = any(sheet.cell(row=row, column=col).value and sheet.cell(row=row, column=col).value.lower() == "mapped" for col in COLS_TO_CHECK_FOR_MAPPING)
            if techname in description_overrides:
                desc = description_overrides[techname]
            if modify.lower() == "n/a":
                modify = None
            if jsonname:
                jsonname = re.sub(r'[^a-zA-Z0-9\[\]\.]', '', jsonname)
            if valuetype == "Enumerated" or valuetype == "Boolean":
                poss_values = ["true", "false"] if valuetype == "Boolean" else typedetails.split(";")
                if example is not None and example not in poss_values:
                    print("WARNING: Invalid example value '{}' for parameter {}".format(example, techname))
                if acadefault is not None and acadefault not in poss_values:
                    print("WARNING: Invalid default value '{}' for parameter {}".format(acadefault, techname))
            if mo.upper() == "M" and example is None:
                print("WARNING: No example value provided for mandatory parameter {}".format(techname))
            if paramtype.lower() in ["input", "return", "inputorreturn"] and cramerstorage not in ["n/a", "{}_GE".format(prodname), "TRANSIENT"] and jsonname is None:
                print("WARNING: No JSON-Name specified for parameter {} with Cramer storage {}".format(techname, cramerstorage.replace("\n", "\\n")))
            if stablenet.lower() not in ["yes", "no"]:
                print("WARNING: Stablenet column for parameter {} is {}. Assuming 'yes'.".format(techname, stablenet))
                stablenet = "yes"
            if paramtype.lower() == "input" or paramtype.lower() == "special":
                DBG(30, "Adding parameter {} (type {}) to input parameters".format(techname, paramtype))
                DBG(50, "valuetype is {}".format(valuetype))
                DBG(50, "acadefault is {}".format(acadefault))
                inparams.append(Param('input', techname, desc, valuetype, typedetails, mo.upper() == "M", example, cramerstorage, acadefault, paramtype.lower() == "special", mapped, maxoccurs, jsonname, modify))
                if techname not in stablenetblacklist:
                    stablenetparams[0].append(techname)
            elif paramtype.lower() == "return":
                DBG(30, "Adding parameter {} (type {}) to cramer output parameters".format(techname, paramtype))
                crameroutparams.append(Param('Cramer', techname, desc, valuetype, typedetails, mo.upper() == "M", example, cramerstorage, maxoccurs=maxoccurs, jsonname=jsonname))
                if techname not in stablenetblacklist:
                    stablenetparams[0].append(techname)
            elif paramtype.lower() == "inputorreturn":
                DBG(30, "Adding parameter {} (type {}) to input AND cramer output parameters".format(techname, paramtype))
                inparams.append(Param('input', techname, desc, valuetype, typedetails, False, example, cramerstorage, acadefault, paramtype.lower() == "special", mapped, maxoccurs, jsonname, modify))
                crameroutparams.append(Param('Cramer', techname, desc, valuetype, typedetails, mo.upper() == "M", example, cramerstorage, maxoccurs=maxoccurs, jsonname=jsonname))
                if techname not in stablenetblacklist:
                    stablenetparams[0].append(techname)
            else:
                DBG(30, "Ignoring parameter {} (type {})".format(techname, paramtype))
            if stablenet.lower() == "yes":
                stablenetparams[1].append(techname)
            if sheet.cell(row=row, column=COLUMNS["TECHNAME"]).font.bold:
                keyparams.append(techname)
        elif techname and not valuetype:
            print("WARNING: No value type defined for parameter '{}'".format(techname))
        elif techname and not mo:
            print("WARNING: Mandatory/Optional column not defined for parameter '{}'".format(techname))
        elif techname and not paramtype:
            print("WARNING: No param type defined for parameter '{}'".format(techname))

    return inparams, crameroutparams, stablenetparams, keyparams

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('-D', dest='dbglevel', action='store', default=10,    help='Print Debug information')
    parser.add_argument('-t', dest='tabname', metavar='<TABNAME>', help='Excel tab read')
    parser.add_argument('-n', dest='nename', metavar='<NENAME_PARAM>', default='NETWORK_ELEMENT_NAME', help='Name of the parameter to be used for network element')
    parser.add_argument('-o', dest='outfile', metavar='<OUTPUT_FILE>', help='Name of the output file (overrides default name)')
    parser.add_argument('filename', metavar='<FILENAME>', help='Excel input file')

    args=parser.parse_args()
    set_debug_level(args.dbglevel)

    prodname, version, inparams, crameroutparams, stablenetparams, keyparams = read_data_from_excel(args.filename, args.tabname)
    config = FactoryProductConfiguration(prodname, version, inparams, crameroutparams, stablenetparams, keyparams)
    if prodname in FP_DEPENDENCIES:
        config.add_prerequisite_product(*FP_DEPENDENCIES[prodname])

    config.add_validation("CHECK_NODE_LOCATION", args.nename, taskname="CHECK_TARGET_NE_EXISTS")
    if "ACCESS_DEVICE_NAME" in config.input_param_names():
        config.add_validation("CHECK_NODE_LOCATION", "ACCESS_DEVICE_NAME", taskname="CHECK_ACCESS_DEVICE_EXISTS")
    outfile = "FP_{}.json".format(prodname) if args.outfile is None else args.outfile
    DBG(10, "Writing json file '{}'".format(outfile))
    with open(outfile, "w", newline='\n') as fp:
        config.to_file(fp)
