#!/usr/bin/python3
import argparse
import re
import sys
import json
from zipfile import ZipFile
from collections import namedtuple
from openpyxl import load_workbook
from lookuptable import LookupTable, LtEntry

from debug import DBG, set_debug_level

def read_data_from_sheet(sheet, col1, col2, cfsname, componentname, fpname):
    STARTROW = 7
    AMBIGUOUS = ["CUST_SNIPPET_NAMES"] # Parameters which can have different values for different FPs

    rv = []
    for row in range(STARTROW, 300):
        cellvalues = [sheet.cell(row=row, column=col).value for col in [1, col1, col2]]
        techname, paramtype, paramdetails = (x.strip() if isinstance(x, str) else x for x in cellvalues)
        if techname == "Version": # avoid reading version history
            break
        if sheet.cell(row=row, column=1).font.strike:
            DBG(20, "Ignoring parameter {} because of strikethough formatting".format(techname))
            continue
        if techname and paramtype and techname != "NETWORK_ELEMENT_NAME":
            if paramtype == 'input':
                if techname in AMBIGUOUS:
                    if cfsname:
                        input_param = "{}_{}".format(fpname, techname)
                    else:
                        input_param = "{}_{}_{}".format(componentname, fpname, techname)
                else:
                    if cfsname:
                        input_param = "{}".format(techname)
                    else:
                        input_param = "{}_{}".format(componentname, techname)
                rv.append({"name": techname, "type": "input", "from": input_param})
            elif paramtype == 'static value':
                rv.append({"name": techname, "type": "static", "value": paramdetails})
            elif paramtype == 'static "null"':
                rv.append({"name": techname, "type": "static", "value": None})
            elif paramtype == 'mapped':
                rv.append({"name": techname, "type": "mapped", "from": paramdetails})
            elif paramtype == 'n/a':
                pass
            else:
                print("Warning: Invalid value '{}' for parameter {}".format(paramtype, techname))
    return rv



def read_data_from_excel(xlfile, cfsname, componentname):
    ''' Read and return the data from the correct sheet in the excel file.'''
    PRODNAME_CELL = 'B1'
    ACTION_CELL = 'B3'
    VERSION_CELL = 'B4'

    HEADERROW = 5 # Row that contains the Component/CFS names

    DBG(10, "Reading data from excel file '{}'".format(xlfile))
    wb = load_workbook(xlfile, data_only=True)
    rv = []
    for tab in wb.sheetnames:
        if re.match(r'[A-Z_]* [A-Z][a-z]*', tab):
            DBG(20, "Checking tab {}".format(tab))
            sheet = wb[tab]

            prodname = sheet[PRODNAME_CELL].value
            if not prodname:
                raise ValueError("Unable to read product name from excel {}, tab {}, cell {}".format(xlfile, tab, PRODNAME_CELL))
            action = sheet[ACTION_CELL].value
            if not action:
                raise ValueError("Unable to read action from excel {}, tab {}, cell {}".format(xlfile, tab, ACTION_CELL))
            version = sheet[VERSION_CELL].value

            for col in range(1,20):
                cellval = sheet.cell(row=HEADERROW, column=col).value
                if  (componentname and cellval == componentname) or (cfsname and cellval == "CFS " + cfsname):
                    col1, col2 = col, col+1
                    DBG(10, "Reading parameters for product {}, action {} from tab {}, columns {},{}".format(prodname, action, tab, col1, col2))
                    data = read_data_from_sheet(sheet, col1, col2, cfsname, componentname, prodname)
                    rv.append({"factoryProduct": prodname, "action": action, "factoryProductVersion": str(version), "parameters": data})
                    break
    
    return rv





if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('-D', dest='dbglevel', action='store', default=10,    help='Print Debug information')
    parser.add_argument('-o', dest='outfile', metavar='<OUTPUT_FILE>', help='Name of the output file (overrides default name)')
    parser.add_argument('filename', metavar='<FILENAME>', help='Excel input file')
    parser.add_argument('composite', metavar='<COMPOSITE>', help='CFS or Component name for which to extract data')

    args=parser.parse_args()
    set_debug_level(args.dbglevel)

    if args.composite.startswith("TN_"):
        cfsname = args.composite
        componentname = None
    else:
        componentname = args.composite
        cfsname = None

    jsondata = {"compositionName": args.composite, "compositionType": "cfs" if cfsname else "component"}
    jsondata["paramMapping"] = read_data_from_excel(args.filename, cfsname, componentname)

    outfile = "pmapping_{}.json".format(args.composite) if args.outfile is None else args.outfile
    DBG(10, "Writing json file '{}'".format(outfile))
    with open(outfile, "w") as fp:
        json.dump(jsondata, fp, indent=2)
