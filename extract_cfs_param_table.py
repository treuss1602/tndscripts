#!/usr/bin/python3
import argparse
import re
import json
from collections import defaultdict
from openpyxl import load_workbook

from debug import DBG, set_debug_level

CFS_COMPONENT_MAPPING = {
    "TN_RBH_RPD_ACCESS": ["RBH_RPD"],
    "TN_RBH_CMTS_ACCESS": ["RBH_CMTS"],
    "TN_RBH_CMTS_CORE": ["RBH_CMTS_INET", "RBH_CMTS_ABR_MC"],
    "TN_B2C_OLT_ACCESS": ["RTL_PROV", "RTL_MGT", "RTL_INET", "RTL_VOIP", "INF_DHCPTRAFFIC", "INF_MGT", "RTL_IPTV", "RTL_CGN"],
    "TN_B2C_XDSLAM_ACCESS": ["RTL_PROV", "RTL_MGT", "RTL_INET", "RTL_VOIP", "INF_DHCPTRAFFIC", "INF_MGT", "RTL_IPTV", "RTL_CGN"],
    "TN_CMTS_ACCESS": ["RBH_CMTS_INET", "RBH_CMTS_ABR_MC"],
    "TN_MBH_MAD": ["MBH_OAM_MAD", "MBH_UPCP_MAD", "MBH_TLMGT_MAD", "MBH_S1_4G_MAD", "MBH_S1_5G_MAD"],
    "TN_MBH_MSA": ["MBH_ACCESS_OAM", "MBH_ACCESS_UPCP", "MBH_ACCESS_TLMGT", "MBH_ACCESS_S1_4G", "MBH_ACCESS_S1_5G"],
    "TN_MBH_4G5G": ["MBH_ACCESS_OAM", "MBH_ACCESS_UPCP", "MBH_ACCESS_S1_4G", "MBH_ACCESS_S1_5G"],
    "TN_MBH_3G": ["MBH_ACCESS_OAM", "MBH_ACCESS_UPCP"],
    "TN_INF_MGT": ["INF_MGT_UT"],
}

COMPATIBILITY_MATRIX = {
    "input" : ["input", "static value", 'static "null"', "mapped", "input with ACA default"],
    "return" : ["n/a"],
    "inputORreturn" : ["input", "static value", 'static "null"', "mapped", "input with ACA default"],
    "n/a" : ["n/a"],
    "special" : ["input"],
    "Stablenet NEI" : ["n/a"]
}

AMBIGUOUS = ["CUST_SNIPPET_NAMES"] # Parameters which can have different values for different FPs

def read_data_from_sheet(sheet, col1, col2, cfsname, componentname, fpname):
    STARTROW = 7

    rv = []
    for row in range(STARTROW, 300):
        cellvalues = [sheet.cell(row=row, column=col).value for col in [1, 3, 4, 11, col1, col2]]
        techname, valuetype, valuedetails, fp_paramtype, paramtype, paramdetails = (x.strip() if isinstance(x, str) else x for x in cellvalues)
        if techname == "Version": # avoid reading version history
            break
        if sheet.cell(row=row, column=1).font.strike:
            DBG(20, "Ignoring parameter {} because of strikethough formatting".format(techname))
            continue
        if techname and paramtype and techname != "NETWORK_ELEMENT_NAME":
            if paramtype not in COMPATIBILITY_MATRIX[fp_paramtype]:
                print("WARNING: {}: type '{}' defined for {} is not compatible with parameter type '{}'".format(techname, paramtype, cfsname if cfsname else componentname, fp_paramtype))
            if paramtype == 'input' or paramtype == 'input with ACA default':
                if techname in AMBIGUOUS:
                    if cfsname:
                        input_param = "{}_{}".format(fpname, techname)
                    else:
                        input_param = "{}_{}_{}".format(componentname, fpname, techname)
                else:
                    if cfsname:
                        input_param = "{}".format(techname)
                    else:
                        input_param = "{}_{}".format(componentname, techname)
                paramentry = {"name": techname, "type": "input", "from": input_param}
                if paramtype == 'input with ACA default':
                    paramentry["acadefault"] = paramdetails
                rv.append(paramentry)
            elif paramtype == 'static value':
                if valuetype == "Enumerated" and paramdetails not in [x.strip() for x in valuedetails.split(";")]:
                    print("Warning: Value '{}' for parameter {} is not a valid enum value.".format(paramdetails, techname))
                if valuetype == "Integer" and not isinstance(paramdetails, int):
                    print("Warning: Value '{}' for parameter {} is not a valid integer.".format(paramdetails, techname))
                rv.append({"name": techname, "type": "static", "value": paramdetails})
            elif paramtype == 'static "null"':
                rv.append({"name": techname, "type": "static", "value": None})
            elif paramtype == 'mapped':
                rv.append({"name": techname, "type": "mapped", "from": paramdetails.replace(" OR ","|").replace(" ","")})
            elif paramtype == 'n/a':
                pass
            else:
                print("Warning: Invalid value '{}' for parameter {}".format(paramtype, techname))
    return rv



def read_data_from_excel(wb, cfsname, componentname):
    ''' Read and return the data from the correct sheet in the excel file.'''
    PRODNAME_CELL = 'B1'
    VERSION_CELL = 'B4'

    HEADERROW = 5 # Row that contains the Component/CFS names

    rv = []
    for tab in wb.sheetnames:
        if re.match(r'[A-Z_]*', tab):
            DBG(20, "Checking tab {}".format(tab))
            sheet = wb[tab]

            prodname = sheet[PRODNAME_CELL].value
            version = sheet[VERSION_CELL].value

            if prodname and version:
                for col in range(1,100):
                    cellval = sheet.cell(row=HEADERROW, column=col).value
                    if  (componentname and cellval == componentname) or (cfsname and cellval == "CFS " + cfsname):
                        col1, col2 = col, col+1
                        DBG(10, "Reading parameters for product {} version {} from tab {}, columns {},{}".format(prodname, version, tab, col1, col2))
                        data = read_data_from_sheet(sheet, col1, col2, cfsname, componentname, prodname)
                        rv.append({"factoryProduct": prodname, "factoryProductVersion": str(version), "parameters": data})
                        break
                    else:
                        if cellval:
                            DBG(50, "Ignoring header {} - not '{}'".format(cellval, "CFS "+cfsname if cfsname else componentname))
            else:
                DBG(20, "Ignoring tab {}".format(tab))

    return rv

def read_component_overrides_from_excel(wb, cfsname):
    overrides = defaultdict(lambda : defaultdict(list))
    sheet = wb["Component Overrides"]
    for row in range(2,101):
        cfs, componentname, fpname, paramname, paramtype, paramdetails = (sheet.cell(row=row, column=c).value for c in range(1,7))
        if not cfs or cfs != cfsname:
            continue
        if not componentname or not fpname or not paramname or not paramtype:
            print("Warning: Row {} in tab 'Component Overrides' is incomplete and thus ignored.".format(row))
            continue
        if paramtype == 'n/a':
            print("Warning: 'n/a' is not applicable as param type in tab 'Component Overrides'. Row {} is ignored.".format(row))
            continue

        if paramtype == 'input' or paramtype == 'input with ACA default':
            if paramname in AMBIGUOUS:
                input_param = "{}_{}_{}".format(componentname, fpname, paramname)
            else:
                input_param = "{}_{}".format(componentname, paramname)
            paramentry = {"name": paramname, "type": "input", "from": input_param}
        elif paramtype == 'static value':
            paramentry = {"name": paramname, "type": "static", "value": paramdetails}
        elif paramtype == 'static "null"':
            paramentry = {"name": paramname, "type": "static", "value": None}
        elif paramtype == 'mapped':
            paramentry = {"name": paramname, "type": "mapped", "from": paramdetails.replace(" OR ","|").replace(" ","")}
        else:
            print("Warning: Invalid param type '{}' in 'Component Overrides', row {}".format(paramtype, row))
            continue
        # Add to complist
        overrides[componentname][fpname].append(paramentry)

    return [{"componentName": comp, 
             "overrides": [{"factoryProduct": fp, "parameters": params} for fp, params in fps.items()]}
            for comp,fps in overrides.items()]


def extract_data_to_file(wb, composite, recursive=False, outfile=None):
    if composite.startswith("TN_"):
        cfsname = composite
        jsondata = {"compositionName": cfsname, "compositionType": "cfs"}
        jsondata["includedComponents"] = CFS_COMPONENT_MAPPING.get(cfsname, [])
        jsondata["paramMapping"] = read_data_from_excel(wb, cfsname, None)
        jsondata["componentOverrides"] = read_component_overrides_from_excel(wb, cfsname)
    else:
        componentname = composite
        jsondata = {"compositionName": componentname, "compositionType": "component"}
        jsondata["paramMapping"] = read_data_from_excel(wb, None, componentname)
        cfsname = None
    
    if outfile is None:
        outfile = "{}_{}.json".format("CFS" if cfsname else "Component", composite)
    DBG(10, "Writing json file '{}'".format(outfile))
    with open(outfile, "w", newline='\n') as fp:
        json.dump(jsondata, fp, indent=2)
    if recursive and cfsname:
        for component in CFS_COMPONENT_MAPPING.get(cfsname, []):
            extract_data_to_file(component)
        


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('-D', dest='dbglevel', action='store', default=10,    help='Print Debug information')
    parser.add_argument('-o', dest='outfile', metavar='<OUTPUT_FILE>', default=None, help='Name of the output file (overrides default name)')
    parser.add_argument('-r', dest='recursive', action="store_true", help='Recurse into creating component jsons if argument is CFS')
    parser.add_argument('filename', metavar='<FILENAME>', help='Excel input file')
    parser.add_argument('composite', metavar='<COMPOSITE>', nargs='+', help='CFS or Component name for which to extract data')

    args=parser.parse_args()
    set_debug_level(args.dbglevel)

    DBG(10, "Reading data from excel file '{}'".format(args.filename))
    wb = load_workbook(args.filename, data_only=True)

    for composite in args.composite:
        extract_data_to_file(wb, composite, args.recursive, args.outfile)
