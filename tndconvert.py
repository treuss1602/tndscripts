#!/usr/bin/python3
import argparse
import sys
import uuid
import validations
from zipfile import ZipFile
from collections import namedtuple
from openpyxl import load_workbook

from factoryproduct import Param, FactoryProductConfiguration
from debug import DBG, set_debug_level

def read_data_from_excel(xlfile, tab):
    ''' Read and return the data from the correct sheet in the excel file.'''
    STARTROW = 6
    COLUMNS = {'TECHNAME': 1, 'DESC': 2, 'VALUETYPE': 3, 'TYPEDETAILS': 4, 'OM': 5, 'EXAMPLE_VALUE': 6, 'PARAMTYPE': 7, 'ACADEFAULT': 9, 'CRAMERSTORAGE': 16}
    PRODNAME_CELL = 'B1'
    ACTION_CELL = 'B3'

    DBG(10, "Reading data from excel file '{}'".format(xlfile))
    wb = load_workbook(xlfile, data_only=True)
    if tab is None:
        sheet = wb.active
        tab = sheet.title
        print("WARNING: No parameter given for <TABNAME>. Reading active tab '{}'".format(tab))
    else:
        DBG(10, "Reading tab '{}'".format(tab))
        try:
            sheet = wb[tab]
        except KeyError:
            tabs = ", ".join("'"+t+"'" for t in wb.sheetnames)
            raise KeyError("Invalid tab name '{}'. Available tabs are: {}".format(tab, tabs))

    prodname = sheet[PRODNAME_CELL].value
    if not prodname:
        raise ValueError("Unable to read product name from excel {}, tab {}, cell {}".format(xlfile, tab, PRODNAME_CELL))
    action = sheet[ACTION_CELL].value
    if not action:
        raise ValueError("Unable to read action from excel {}, tab {}, cell {}".format(xlfile, tab, ACTION_CELL))
    inparams = []
    crameroutparams = []
    for row in range(STARTROW, 300):
        cellvalues = [sheet.cell(row=row, column=col).value for col in [COLUMNS[s]
                                                     for s in ['TECHNAME','DESC','VALUETYPE','TYPEDETAILS','OM','EXAMPLE_VALUE','PARAMTYPE', 'ACADEFAULT', 'CRAMERSTORAGE']]]
        techname, desc, valuetype, typedetails, mo, example, paramtype, acadefault, cramerstorage = (x.strip() if isinstance(x, str) else None for x in cellvalues)
        if techname and valuetype and mo and paramtype:
            if sheet.cell(row=row, column=1).font.strike:
                DBG(10, "Ignoring parameter {} because of strikethough formatting".format(techname))
                continue
            if paramtype.lower() == "input" or paramtype.lower() == "special":
                DBG(30, "Adding parameter {} (type {}) to input parameters".format(techname, paramtype))
                inparams.append(Param('input', techname, desc, valuetype, typedetails, mo.upper() == "M", example, cramerstorage, acadefault, paramtype.lower() == "special"))
            elif paramtype.lower() == "return":
                crameroutparams.append(Param('Cramer', techname, desc, valuetype, typedetails, mo.upper() == "M", example, cramerstorage))
                DBG(30, "Adding parameter {} (type {}) to input cramer output parameters".format(techname, paramtype))
            else:
                DBG(30, "Ignoring parameter {} (type {})".format(techname, paramtype))

    return prodname, action, inparams, crameroutparams

def create_zipfile(filename, *tables):
    DBG(10, "Writing Lookup Tables zipfile '{}'".format(filename))
    with ZipFile(filename, "w") as z:
        for table in tables:
            z.writestr(table.name, table.dump())

def create_sample_order(productconfig : FactoryProductConfiguration, orderno = None, replyto_address = None, mandatory_only = True):
    if orderno is None:
        orderno = uuid.uuid4()
    xml = """<?xml version="1.0"?>
<soap:Envelope xmlns:soap="http://www.w3.org/2003/05/soap-envelope"
               xmlns:il="http://soa.comptel.com/2011/02/instantlink">
   <soap:Header xmlns:wsa="http://www.w3.org/2005/08/addressing">
"""
    if replyto_address is not None:
        xml += """      <wsa:ReplyTo>
         <wsa:Address>{}</wsa:Address>
      </wsa:ReplyTo>
""".format(replyto_address)
    xml += """   </soap:Header>
   <soap:Body>
      <il:CreateRequest>
         <il:RequestHeader>
            <il:NeType>CDFF</il:NeType>
            <il:OrderNo>{}</il:OrderNo>
         </il:RequestHeader>
         <il:RequestParameters>
""".format(orderno)
    xml += """            <il:Parameter name="IL_REQ_GROUP" value="TND" />
            <il:Parameter name="ORDER_TYPE" value="Connect" />
            <il:Parameter name="ORDER_EXTERNAL_ORDER_ID" value="{}" />
            <!-- Order Line 1: Create FACTORY_PRODUCT -->
            <il:Parameter name="LINE_1_ORDER_LINE_ID" value="OL_1" />
            <il:Parameter name="LINE_1_NAME" value="FACTORY_PRODUCT_{}" />
            <il:Parameter name="LINE_1_ACTION" value="Create" />
""".format(orderno, productconfig.factoryProductName)
    for p in productconfig.input_params:
        if p.mandatory or not mandatory_only:
            xml += '            <il:Parameter name="LINE_1_PS_PARAM_{}" value="{}" />\n'.format(p.name, p.examplevalue)
    xml += """         </il:RequestParameters>
      </il:CreateRequest>
   </soap:Body>
</soap:Envelope>"""
    return xml

def create_confluence_table(productconfig : FactoryProductConfiguration):
    print("h2. Factory Product {}".format(productconfig.factoryProductName))
    print("h3. {}".format(productconfig.transaction))
    print("||Parameter name (without prefix)||Section||Description||Example Value||M/O/C||")
    print("|EXTERNAL_ORDER_ID|General|Reference to the order number. Generated by the northbound system.|O00324572912|M|")
    print("|CONTEXT_ID|Order Line|Cramer context which to relate the project for creation of this factory product.\nNormally not used.|Create-PHY_SINGLE_LINK-O00324572912|O|")
    for param in productconfig.input_params:
        print("|{}|Order Line|{}|{}|{}|".format(param.name, param.desc, param.examplevalue, "M" if param.mandatory else "O"))

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('-D', dest='dbglevel', action='store', default=10,    help='Print Debug information')
    parser.add_argument('-t', dest='tabname', metavar='<TABNAME>', help='Excel tab read')
    parser.add_argument('-n', dest='nename', metavar='<NENAME_PARAM>', default='NETWORK_ELEMENT_NAME', help='Name of the parameter to be used for network element')
    parser.add_argument('-o', dest='outfile', metavar='<OUTPUT_FILE>', help='Name of the output file (overrides default name)')
    parser.add_argument('-m', dest='mandatory_only', action='store_true', help='Create request with mandatory parameters only (applies only to j2r action)')
    parser.add_argument('filename', metavar='<FILENAME>', help='Input file (Excel or Json)')
    parser.add_argument('action', metavar='<ACTION>', choices=["x2j", "j2t", "x2t", "j2r", "j2c"], nargs="?", help='Conversion to apply: Excel to json (x2j), JSON to Lookup Tables (j2t) or Exel to Lookup Tables (x2t)')

    args=parser.parse_args()
    set_debug_level(args.dbglevel)

    if args.action is None and args.filename.endswith(".xlsx"):
        action = "x2j"
    elif args.action is None and args.filename.endswith(".json"):
        action = "j2t"
    elif args.action is None:
        print("ERROR: Unable to derive action from input filename and no argument given.")
        sys.exit(1)
    else:
        action = args.action

    if action == "x2j":
        prodname, transaction, inparams, crameroutparams = read_data_from_excel(args.filename, args.tabname)
        config = FactoryProductConfiguration(prodname, transaction, inparams, crameroutparams)
        config.add_validation("CHECK_NODE_LOCATION", args.nename, taskname="CHECK_TARGET_NE_EXISTS")
        if "ACCESS_DEVICE_NAME" in config.input_param_names():
            #config.add_validation("CHECK_NODE_LOCATION", "ACCESS_DEVICE_NAME", taskname="CHECK_ACCESS_DEVICE_EXISTS")
            config.add_validation("CHECK_NODE_LOCATION", "ACCESS_DEVICE_NAME")
        outfile = "{}_{}.json".format(prodname, transaction) if args.outfile is None else args.outfile
        DBG(10, "Writing json file '{}'".format(outfile))
        with open(outfile, "w") as fp:
            config.to_file(fp)
    elif action == "j2t":
        DBG(10, "Reading json file '{}'".format(args.filename))
        with open(args.filename, "r") as fp:
            config = FactoryProductConfiguration.from_file(fp)
        tables = config.create_lookup_tables(args.nename)
        for table in tables:
            DBG(30, "Lookup Table dump:\n"+table.debugdump())
        outfile = "{}_{}.zip".format(config.factoryProductName, config.transaction) if args.outfile is None else args.outfile
        create_zipfile(outfile, *tables)
    elif action == "x2t":
        prodname, transaction, inparams, crameroutparams = read_data_from_excel(args.filename, args.tabname)
        validations_ = [validations.NodeExistsValidation(args.nename)] # Default validation of NETWORK_ELEMENT_NAME
        config = FactoryProductConfiguration(prodname, transaction, inparams, crameroutparams, validations_)
        tables = config.create_lookup_tables(args.nename)
        for table in tables:
            DBG(30, "Lookup Table dump:\n"+table.debugdump())
        outfile = "{}_{}.zip".format(config.factoryProductName, config.transaction) if args.outfile is None else args.outfile
        create_zipfile(outfile, *tables)
    elif action == "j2r":
        DBG(10, "Reading json file '{}'".format(args.filename))
        with open(args.filename, "r") as fp:
            config = FactoryProductConfiguration.from_file(fp)
        req = create_sample_order(config, mandatory_only=args.mandatory_only)
        print(req)
    elif action == "j2c":
        DBG(10, "Reading json file '{}'".format(args.filename))
        with open(args.filename, "r") as fp:
            config = FactoryProductConfiguration.from_file(fp)
        create_confluence_table(config)
        #print(req)
    else:
        print("ERROR: Unknown action value '{}'".format(action))
        sys.exit(1)
